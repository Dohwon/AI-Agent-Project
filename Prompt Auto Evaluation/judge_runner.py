r"""
Judge Runner: N-way round-robin (pairwise) model comparison with LLM-as-Judge.
- Reads tests from CSV (extended schema)
- Calls ALL models (MODELS=A,B[,C...]) concurrently per query (one generation per model)
- Runs pairwise judging for all combinations (round-robin), aggregates wins/points
- Preserves legacy outputs for the primary pair (first two tags in MODELS):
  * model_a_output/model_b_output/evaluation/usage_breakdown keys remain
- Writes per-query JSONL records (backward compatible + extra fields)
- Builds an Excel workbook (same logic; extra fields appear as extra columns)
No pandas required. Uses httpx + openpyxl only.
"""

import asyncio
import csv
import json
import os
import time
from contextlib import AsyncExitStack
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import httpx
from dotenv import load_dotenv
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Py3.11/3.12 호환용 UTC 상수
try:
    from datetime import UTC
except Exception:
    from datetime import timezone as _tz
    UTC = _tz.utc


# -------------------------
# Helpers
# -------------------------

def _coerce_float(x: Optional[str]) -> Optional[float]:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None

def _normalize_path(p: str) -> str:
    # Windows-style ".\foo.txt" or ".\.\foo.txt"를 리눅스/맥에서도 읽기 좋게 정규화
    s = (p or "").strip().strip('"').strip("'")
    s = s.replace("\\", "/")
    if s.startswith("././"):
        s = "./" + s[4:]
    return s

def _read_text_file(path_str: str) -> str:
    p = Path(_normalize_path(path_str))
    return p.read_text(encoding="utf-8")

def _uniquify_path(p: Path) -> Path:
    """이미 존재하면 ' (1)', ' (2)' … 붙여서 충돌 나지 않는 경로를 돌려준다."""
    if not p.exists():
        return p
    i = 1
    stem, suffix = p.stem, p.suffix
    while True:
        cand = p.with_name(f"{stem} ({i}){suffix}")
        if not cand.exists():
            return cand
        i += 1

def make_dated_output_paths(out_dir: Path, base_prefix: str = "results") -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    ymd = datetime.now().strftime("%y%m%d")
    jsonl = _uniquify_path(out_dir / f"{base_prefix}_{ymd}.jsonl")
    xlsx  = _uniquify_path(out_dir / f"{base_prefix}_{ymd}.xlsx")
    return jsonl, xlsx


# -------------------------
# ENV / Config
# -------------------------

def _parse_models_list(s: str) -> List[str]:
    raw = (s or "").strip()
    if not raw:
        return ["A", "B"]
    return [t.strip() for t in raw.split(",") if t.strip()]

def load_env() -> Dict[str, Any]:
    load_dotenv()

    models = _parse_models_list(os.getenv("MODELS", "A,B"))

    model_cfg: Dict[str, Dict[str, Any]] = {}
    costs: Dict[str, Optional[float]] = {}

    for tag in models:
        pfx = f"MODEL_{tag}_"
        vendor = (os.getenv(pfx + "VENDOR", "openai") or "openai").strip().lower()
        base_url = (os.getenv(pfx + "BASE_URL", "") or "").strip()
        api_key = (os.getenv(pfx + "API_KEY", "") or "").strip()
        model = (os.getenv(pfx + "MODEL", "") or "").strip()
        sys_prompt_path = (os.getenv(pfx + "SYS_PROMPT", "") or "").strip()

        if not base_url or not api_key or not model:
            raise ValueError(f"[ENV] Missing required fields for {tag}: BASE_URL/API_KEY/MODEL")

        model_cfg[tag] = {
            "tag": tag,
            "vendor": vendor,
            "base_url": base_url,
            "api_key": api_key,
            "model": model,
            "sys_prompt_path": sys_prompt_path,
        }

        costs[f"{tag}_IN"] = _coerce_float(os.getenv(pfx + "INPUT_COST_PER_1K"))
        costs[f"{tag}_OUT"] = _coerce_float(os.getenv(pfx + "OUTPUT_COST_PER_1K"))
        costs[f"{tag}_CACHED"] = _coerce_float(os.getenv(pfx + "CACHED_COST_PER_1K"))

    judge = {
        "vendor": (os.getenv("JUDGE_VENDOR", "openai") or "openai").strip().lower(),
        "base_url": (os.getenv("JUDGE_BASE_URL", "") or "").strip(),
        "api_key": (os.getenv("JUDGE_API_KEY", "") or "").strip(),
        "model": (os.getenv("JUDGE_MODEL", "") or "").strip(),
    }
    if not judge["base_url"] or not judge["api_key"] or not judge["model"]:
        raise ValueError("[ENV] Missing required fields for Judge: JUDGE_BASE_URL/JUDGE_API_KEY/JUDGE_MODEL")

    costs["J_IN"] = _coerce_float(os.getenv("JUDGE_INPUT_COST_PER_1K"))
    costs["J_OUT"] = _coerce_float(os.getenv("JUDGE_OUTPUT_COST_PER_1K"))
    costs["J_CACHED"] = _coerce_float(os.getenv("JUDGE_CACHED_COST_PER_1K"))

    cfg = {
        "MODELS": models,
        "MODEL_CFG": model_cfg,
        "JUDGE": judge,
        "CONC": int(os.getenv("MAX_CONCURRENCY", "4")),
        "TIMEOUT": int(os.getenv("TIMEOUT", "90")),
        "COST": costs,
    }
    return cfg


# -------------------------
# HTTP / Vendors
# -------------------------

def make_client(vendor: str, base_url: str, api_key: str, timeout: int) -> httpx.AsyncClient:
    vendor = (vendor or "openai").lower()
    headers: Dict[str, str] = {}
    if vendor == "gemini":
        headers = {"x-goog-api-key": api_key}
    else:
        headers = {"Authorization": f"Bearer {api_key}"}
    return httpx.AsyncClient(timeout=timeout, headers=headers)

def make_chat_url(base_url: str) -> str:
    base = base_url.rstrip("/")
    if base.endswith("/v1/chat/completions"):
        return base
    if base.endswith("/v1"):
        return base + "/chat/completions"
    return base + "/v1/chat/completions"

async def call_chat(
    client: httpx.AsyncClient,
    base_url: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    force_json: bool = False,
) -> Tuple[bool, Dict[str, Any] or str, Dict[str, Any]]:
    url = make_chat_url(base_url)
    started = time.perf_counter()
    try:
        body = {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "stream": False,
        }
        if force_json:
            body["response_format"] = {"type": "json_object"}

        resp = await client.post(url, json=body)
        resp.raise_for_status()
        data = resp.json()

        try:
            content = data["choices"][0]["message"]["content"]
        except Exception:
            content = json.dumps(data, ensure_ascii=False)

        usage = data.get("usage") or {}
        in_tok = usage.get("prompt_tokens") or usage.get("input_tokens")
        out_tok = usage.get("completion_tokens") or usage.get("output_tokens")
        tot_tok = usage.get("total_tokens")
        cached_tok = usage.get("cached_tokens")

        if tot_tok is None and (in_tok is not None or out_tok is not None):
            try:
                tot_tok = (in_tok or 0) + (out_tok or 0)
            except Exception:
                tot_tok = None

        meta = {
            "duration": round(time.perf_counter() - started, 4),
            "usage": {
                "input_tokens": in_tok,
                "output_tokens": out_tok,
                "total_tokens": tot_tok,
                "cached_tokens": cached_tok,
            }
        }
        return True, {"role": "assistant", "content": content}, meta

    except httpx.HTTPError as e:
        meta = {"duration": round(time.perf_counter() - started, 4), "usage": {}}
        return False, f"HTTPError: {str(e)}", meta
    except Exception as e:
        meta = {"duration": round(time.perf_counter() - started, 4), "usage": {}}
        return False, f"Error: {repr(e)}", meta


def to_gemini_request(messages, temperature=0.7, top_p=0.95, max_tokens=1024):
    system_txt = "\n".join(m.get("content", "") for m in messages if m.get("role") == "system")

    contents = []
    for m in messages:
        role = m.get("role")
        if role == "system":
            continue
        gem_role = "user" if role == "user" else "model"
        contents.append({
            "role": gem_role,
            "parts": [{"text": m.get("content", "")}]
        })

    body = {
        "contents": contents,
        "generationConfig": {
            "temperature": temperature,
            "topP": top_p,
            "maxOutputTokens": max_tokens
        },
        "safetySettings": [
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH",       "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HARASSMENT",        "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
        ]
    }
    if system_txt.strip():
        body["systemInstruction"] = {"parts": [{"text": system_txt}]}
    return body


async def call_gemini_chat(
    client: httpx.AsyncClient,
    base_url: str,
    model: str,
    api_key: str,
    system_prompt: str,
    user_prompt: str,
    force_json: bool = False,
) -> Tuple[bool, Dict[str, Any] or str, Dict[str, Any]]:
    url = f"{base_url.rstrip('/')}/v1beta/models/{model}:generateContent"
    started = time.perf_counter()
    try:
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_prompt},
        ]
        payload = to_gemini_request(messages, temperature=0.7, top_p=0.95, max_tokens=1024)
        if force_json:
            payload.setdefault("generationConfig", {})["responseMimeType"] = "application/json"

        headers = {"Content-Type": "application/json", "x-goog-api-key": api_key}
        resp = await client.post(url, headers=headers, json=payload)
        resp.raise_for_status()
        data = resp.json()

        try:
            parts = data["candidates"][0]["content"]["parts"]
            content = "".join(p.get("text", "") for p in parts)
        except Exception:
            content = json.dumps(data, ensure_ascii=False)

        um = data.get("usageMetadata") or {}
        meta = {
            "duration": round(time.perf_counter() - started, 4),
            "usage": {
                "input_tokens":  um.get("promptTokenCount"),
                "output_tokens": um.get("candidatesTokenCount"),
                "total_tokens":  um.get("totalTokenCount"),
            }
        }
        return True, {"role": "assistant", "content": content}, meta

    except httpx.HTTPError as e:
        meta = {"duration": round(time.perf_counter() - started, 4), "usage": {}}
        return False, f"HTTPError: {str(e)}", meta
    except Exception as e:
        meta = {"duration": round(time.perf_counter() - started, 4), "usage": {}}
        return False, f"Error: {repr(e)}", meta


async def call_model_chat(
    vendor: str,
    client: httpx.AsyncClient,
    base_url: str,
    api_key: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    force_json: bool = False,
) -> Tuple[bool, Dict[str, Any] or str, Dict[str, Any]]:
    vendor = (vendor or "openai").lower()
    if vendor == "gemini":
        return await call_gemini_chat(client, base_url, model, api_key, system_prompt, user_prompt, force_json=force_json)
    return await call_chat(client, base_url, model, system_prompt, user_prompt, force_json=force_json)


# -------------------------
# Parsing / Validation
# -------------------------

def try_parse_json_object(text: Optional[str]) -> Optional[dict]:
    if not text:
        return None
    t = text.strip()
    try:
        return json.loads(t)
    except Exception:
        pass
    try:
        s = t.find("{")
        e = t.rfind("}")
        if s != -1 and e != -1 and e > s:
            return json.loads(t[s:e+1])
    except Exception:
        return None
    return None

def validate_case2(payload: dict, func_spec: Optional[dict]) -> Dict[str, Any]:
    out = {
        "function_name": None,
        "available_function": None,
        "function_in_whitelist": None,
        "available_in_whitelist": None,
    }
    if not isinstance(payload, dict) or payload.get("case") != 2:
        return out
    fn = payload.get("function")
    af = payload.get("available_function")
    out["function_name"] = fn
    out["available_function"] = af
    if func_spec:
        out["function_in_whitelist"] = fn in func_spec
        if out["function_in_whitelist"]:
            out["available_in_whitelist"] = (af in func_spec.get(fn, []))
    return out

def validate_case3(payload: dict, min_id: Optional[int], max_id: Optional[int]) -> Dict[str, Any]:
    out = {"prepared_number": None, "prepared_in_range": None}
    if not isinstance(payload, dict) or payload.get("case") != 3:
        return out
    val = payload.get("prepared_question_list_number")
    n = None
    try:
        if val is not None:
            n = int(str(val).strip())
    except Exception:
        n = None
    out["prepared_number"] = val
    if n is not None and min_id is not None and max_id is not None:
        out["prepared_in_range"] = (min_id <= n <= max_id)
    return out


# -------------------------
# Judge payload
# -------------------------

def build_judge_user_content(
    user_input: str,
    expected_case: Any,
    a_label: str,
    a_output: str,
    b_label: str,
    b_output: str
) -> str:
    return (
        "## Inputs\n"
        f"- User query: {user_input}\n"
        f"- Expected case: {expected_case}\n\n"
        f"### Response A ({a_label})\n"
        f"{a_output}\n\n"
        f"### Response B ({b_label})\n"
        f"{b_output}\n\n"
        "## Task\n"
        "Evaluate per the system instructions and output JSON ONLY."
    )


# -------------------------
# Excel writer (원본 로직 유지)
# -------------------------

def flatten(obj: Any, prefix: str = "", out: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    if out is None:
        out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            flatten(v, key, out)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            key = f"{prefix}[{i}]"
            flatten(v, key, out)
    else:
        out[prefix] = obj
    return out

def top_group(col: str) -> str:
    return col.split(".", 1)[0] if "." in col else col

def write_excel_from_jsonl(jsonl_path: Path, xlsx_path: Path, summary: Dict[str, Any]) -> None:
    records: List[dict] = []
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except Exception:
                continue

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    def write_kv(start_row: int, title: str, kv: Dict[str, Any]) -> int:
        ws_sum.cell(row=start_row, column=1, value=title).font = Font(bold=True)
        r = start_row + 1
        for k, v in kv.items():
            ws_sum.cell(row=r, column=1, value=str(k))
            ws_sum.cell(row=r, column=2, value=str(v))
            r += 1
        ws_sum.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        return r + 1

    row = 1
    scalar_top = {
        "total_evaluations": summary.get("total_evaluations"),
        "successful_evaluations": summary.get("successful_evaluations"),
        "failed_evaluations": summary.get("failed_evaluations"),
        "timestamp": summary.get("timestamp"),
    }
    row = write_kv(row, "Overview", scalar_top)
    row = write_kv(row, "Overall Results", summary.get("overall_results", {}))
    row = write_kv(row, "By Case", {k: v for k, v in summary.get("by_case", {}).items()})
    row = write_kv(row, "Confidence Distribution", summary.get("confidence_distribution", {}))
    row = write_kv(row, "Token Usage (Total & By Model)", summary.get("token_usage", {}))
    row = write_kv(row, "Cost Estimate (By Model)", summary.get("cost_estimate", {}).get("by_model", {}))
    row = write_kv(row, "Performance", summary.get("performance", {}))

    # ✅ 추가 섹션(기존 유지 + 확장)
    if summary.get("tournament_results"):
        row = write_kv(row, "Tournament (Round Robin) Results", summary.get("tournament_results", {}))

    for c in (1, 2):
        ws_sum.column_dimensions[get_column_letter(c)].width = 40

    ws = wb.create_sheet("Results")

    flat_rows: List[Dict[str, Any]] = []
    cols_set = set()
    for rec in records:
        fr = flatten(rec)
        flat_rows.append(fr)
        cols_set.update(fr.keys())

    pinned = ["user_input", "query_index", "expected_case"]
    pinned_present = [c for c in pinned if c in cols_set]
    for c in pinned_present:
        cols_set.discard(c)

    rest_cols = sorted(cols_set)
    cols = pinned_present + rest_cols

    ws.append(["" for _ in cols])
    ws.append(cols)

    group_to_cols: Dict[str, List[int]] = {}
    for idx, col in enumerate(cols, start=1):
        g = top_group(col)
        group_to_cols.setdefault(g, []).append(idx)

    for g, idxs in group_to_cols.items():
        start = min(idxs)
        end = max(idxs)
        if g in pinned_present:
            ws.cell(row=1, column=start, value=g)
            ws.merge_cells(start_row=1, start_column=start, end_row=2, end_column=start)
        else:
            ws.cell(row=1, column=start, value=g)
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    for r in (1, 2):
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

    for fr in flat_rows:
        ws.append([fr.get(k, "") for k in cols])

    for i, col in enumerate(cols, start=1):
        max_len = max(len(str(col)), *(len(str(r.get(col, ""))) for r in flat_rows))
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 70)

    try:
        wb.save(xlsx_path)
    except PermissionError:
        xlsx_path = _uniquify_path(xlsx_path)
        wb.save(xlsx_path)


# -------------------------
# Summary builder (확장 + 기존 유지)
# -------------------------

def build_summary(records: List[dict], costs: Dict[str, Optional[float]], model_tags: List[str]) -> Dict[str, Any]:
    total = len(records)
    successes = sum(1 for r in records if r.get("evaluation_status") == "success")
    fails = total - successes

    # legacy: primary pair A/B win counts (evaluation.winner)
    model_a_wins = model_b_wins = ties = 0
    by_case: Dict[str, Dict[str, int]] = {}
    conf_dist = {f"confidence_{i}": 0 for i in range(1, 6)}

    total_input_tok = total_output_tok = total_tok = 0
    total_dur = 0.0

    # by-tag aggregation (새)
    agg_by_tag: Dict[str, Dict[str, Any]] = {}
    def ensure_tag(t: str):
        if t not in agg_by_tag:
            agg_by_tag[t] = {"in": 0, "out": 0, "tot": 0, "cached": 0, "dur": 0.0}

    # tournament aggregation (새)
    tour_points: Dict[str, float] = {t: 0.0 for t in model_tags}
    tour_champ_counts: Dict[str, int] = {t: 0 for t in model_tags}

    for rec in records:
        ev = rec.get("evaluation") or {}
        winner = ev.get("winner")
        if winner == "A":
            model_a_wins += 1
        elif winner == "B":
            model_b_wins += 1
        elif winner == "tie":
            ties += 1

        exp_case = str(rec.get("expected_case"))
        by_case.setdefault(f"case_{exp_case}", {"a_wins": 0, "b_wins": 0, "ties": 0, "total": 0})
        by_case[f"case_{exp_case}"]["total"] += 1
        if winner == "A":
            by_case[f"case_{exp_case}"]["a_wins"] += 1
        elif winner == "B":
            by_case[f"case_{exp_case}"]["b_wins"] += 1
        elif winner == "tie":
            by_case[f"case_{exp_case}"]["ties"] += 1

        conf = ev.get("confidence")
        if isinstance(conf, int) and 1 <= conf <= 5:
            conf_dist[f"confidence_{conf}"] += 1

        meta = rec.get("api_call_metadata") or {}
        total_dur += float(meta.get("duration_seconds") or 0.0)
        total_input_tok += int(meta.get("input_tokens") or 0)
        total_output_tok += int(meta.get("output_tokens") or 0)
        total_tok += int(meta.get("total_tokens") or 0)

        # ✅ 확장: usage_breakdown_by_tag 우선 사용, 없으면 legacy usage_breakdown 사용
        ubt = rec.get("usage_breakdown_by_tag")
        if isinstance(ubt, dict) and ubt:
            for t, d in ubt.items():
                ensure_tag(t)
                agg_by_tag[t]["in"] += int(d.get("input_tokens") or 0)
                agg_by_tag[t]["out"] += int(d.get("output_tokens") or 0)
                agg_by_tag[t]["tot"] += int(d.get("total_tokens") or 0)
                agg_by_tag[t]["cached"] += int(d.get("cached_tokens") or 0)
                agg_by_tag[t]["dur"] += float(d.get("duration") or 0.0)
        else:
            ub = rec.get("usage_breakdown") or {}
            for t in ("A", "B", "J"):
                d = ub.get(t) or {}
                ensure_tag(t)
                agg_by_tag[t]["in"] += int(d.get("input_tokens") or 0)
                agg_by_tag[t]["out"] += int(d.get("output_tokens") or 0)
                agg_by_tag[t]["tot"] += int(d.get("total_tokens") or 0)
                agg_by_tag[t]["cached"] += int(d.get("cached_tokens") or 0)
                agg_by_tag[t]["dur"] += float(d.get("duration") or 0.0)

        # tournament
        tinfo = rec.get("tournament") or {}
        sc = tinfo.get("scoreboard") or {}
        for t in model_tags:
            if t in sc:
                try:
                    tour_points[t] += float(sc[t].get("points") or 0.0)
                except Exception:
                    pass
        champ = (tinfo.get("champion") or {}).get("tag")
        if champ in tour_champ_counts:
            tour_champ_counts[champ] += 1

    def rate(part, whole): 
        return f"{(part / whole * 100):.1f}%" if whole else "0.0%"

    overall_results = {
        # legacy (primary pair)
        "model_a_wins": model_a_wins,
        "model_b_wins": model_b_wins,
        "ties": ties,
        "model_a_win_rate": rate(model_a_wins, total),
        "model_b_win_rate": rate(model_b_wins, total),
        "tie_rate": rate(ties, total),
    }

    token_usage = {
        "total_input_tokens": total_input_tok,
        "total_output_tokens": total_output_tok,
        "total_tokens": total_tok,
        # 확장
        "by_tag": {
            t: {"input": v["in"], "output": v["out"], "total": v["tot"], "cached": v["cached"]}
            for t, v in agg_by_tag.items()
        }
    }

    def calc(tokens: int, per_1k: Optional[float]) -> Optional[float]:
        if per_1k is None:
            return None
        return (tokens / 1000.0) * per_1k

    cost_by_model: Dict[str, Any] = {}
    for t, v in agg_by_tag.items():
        in_cost = calc(v["in"], costs.get(f"{t}_IN"))
        out_cost = calc(v["out"], costs.get(f"{t}_OUT"))
        cached_cost = calc(v["cached"], costs.get(f"{t}_CACHED")) if costs.get(f"{t}_CACHED") is not None else None
        total_cost = None
        if in_cost is not None and out_cost is not None:
            total_cost = in_cost + out_cost + (cached_cost or 0)
        cost_by_model[t] = {
            "input_cost_usd":  f"${in_cost:.2f}" if in_cost is not None else "N/A",
            "output_cost_usd": f"${out_cost:.2f}" if out_cost is not None else "N/A",
            "cached_cost_usd": f"${cached_cost:.2f}" if cached_cost is not None else "N/A",
            "total_cost_usd":  f"${total_cost:.2f}" if total_cost is not None else "N/A",
        }

    performance = {
        "total_duration_seconds": round(total_dur, 2),
        "average_duration_per_query": round(total_dur / total, 2) if total else 0.0,
    }

    # tournament summary
    champion_overall = None
    if tour_points:
        champion_overall = max(tour_points.items(), key=lambda kv: kv[1])[0]

    tournament_results = {
        "models": ",".join(model_tags),
        "points_by_tag": {k: round(v, 2) for k, v in tour_points.items()},
        "champions_by_query": tour_champ_counts,
        "overall_champion_by_points": champion_overall,
    }

    return {
        "total_evaluations": total,
        "successful_evaluations": successes,
        "failed_evaluations": fails,
        "overall_results": overall_results,
        "by_case": by_case,
        "confidence_distribution": conf_dist,
        "token_usage": token_usage,
        "cost_estimate": {"by_model": cost_by_model},
        "performance": performance,
        "tournament_results": tournament_results,
        "timestamp": datetime.now(UTC).isoformat()
    }


# -------------------------
# Core runner
# -------------------------

def _tok(meta: Dict[str, Any], key: str) -> int:
    try:
        return int(((meta.get("usage") or {}).get(key)) or 0)
    except Exception:
        return 0

def _dur(meta: Dict[str, Any]) -> float:
    try:
        return float(meta.get("duration") or 0.0)
    except Exception:
        return 0.0

def _eq(a, b):
    return (str(a).strip() == str(b).strip()) if (a is not None and b is not None) else None


async def process_one(
    idx: int,
    row: Dict[str, str],
    cfg: Dict[str, Any],
    clients: Dict[str, httpx.AsyncClient],
    sys_prompts: Dict[str, str],
    judge_sys: str,
    func_spec: Optional[dict],
    prepared_min: Optional[int],
    prepared_max: Optional[int],
    sema: asyncio.Semaphore,
    pair_pbar: Optional[tqdm] = None,
) -> dict:
    query = (row.get("query") or "").strip()
    exp_case = row.get("expected_case")
    try:
        exp_case = int(exp_case) if exp_case not in (None, "") else None
    except Exception:
        exp_case = None

    exp_fn  = (row.get("expected_function") or "").strip() or None
    exp_act = (row.get("expected_available_function") or "").strip() or None
    exp_pq  = (row.get("expected_prepared_question") or "").strip() or None

    model_tags: List[str] = cfg["MODELS"]
    if len(model_tags) < 2:
        raise ValueError("[RUN] MODELS must include at least 2 tags, e.g., MODELS=A,B")

    # primary pair = first two tags in MODELS (legacy A/B slots)
    tagA = model_tags[0]
    tagB = model_tags[1]

    async def limited_call(coro):
        async with sema:
            return await coro

    # 1) Generate model outputs (one per tag)
    model_results: Dict[str, Tuple[bool, Any, Dict[str, Any]]] = {}
    gen_tasks = {}
    for t in model_tags:
        mc = cfg["MODEL_CFG"][t]
        gen_tasks[t] = asyncio.create_task(
            limited_call(
                call_model_chat(
                    vendor=mc["vendor"],
                    client=clients[t],
                    base_url=mc["base_url"],
                    api_key=mc["api_key"],
                    model=mc["model"],
                    system_prompt=sys_prompts[t],
                    user_prompt=query,
                    force_json=False
                )
            )
        )
    for t, task in gen_tasks.items():
        model_results[t] = await task

    def content_of(t: str) -> str:
        ok, payload, _meta = model_results[t]
        if ok:
            return payload.get("content") or ""
        return ""

    a_ok, a_payload, a_meta = model_results[tagA]
    b_ok, b_payload, b_meta = model_results[tagB]
    a_content = a_payload["content"] if a_ok else ""
    b_content = b_payload["content"] if b_ok else ""

    # 2) Parse/validate primary A/B (legacy fields)
    a_parsed = try_parse_json_object(a_content) if a_ok else None
    b_parsed = try_parse_json_object(b_content) if b_ok else None

    a_case = (a_parsed or {}).get("case")
    b_case = (b_parsed or {}).get("case")

    a_c2 = validate_case2(a_parsed or {}, func_spec)
    b_c2 = validate_case2(b_parsed or {}, func_spec)
    a_c3 = validate_case3(a_parsed or {}, prepared_min, prepared_max)
    b_c3 = validate_case3(b_parsed or {}, prepared_min, prepared_max)

    a_func_eq = _eq((a_parsed or {}).get("function"), exp_fn) if a_case == 2 and exp_fn else None
    a_act_eq  = _eq((a_parsed or {}).get("available_function"), exp_act) if a_case == 2 and exp_act else None
    b_func_eq = _eq((b_parsed or {}).get("function"), exp_fn) if b_case == 2 and exp_fn else None
    b_act_eq  = _eq((b_parsed or {}).get("available_function"), exp_act) if b_case == 2 and exp_act else None

    a_pq_eq   = _eq((a_parsed or {}).get("prepared_question_list_number"), exp_pq) if a_case == 3 and exp_pq else None
    b_pq_eq   = _eq((b_parsed or {}).get("prepared_question_list_number"), exp_pq) if b_case == 3 and exp_pq else None

    # 3) Round-robin pairwise judging
    pairs: List[Tuple[str, str]] = []
    for i in range(len(model_tags)):
        for j in range(i + 1, len(model_tags)):
            pairs.append((model_tags[i], model_tags[j]))

    judge_vendor = cfg["JUDGE"]["vendor"]
    judge_base = cfg["JUDGE"]["base_url"]
    judge_key = cfg["JUDGE"]["api_key"]
    judge_model = cfg["JUDGE"]["model"]

    async def judge_pair(left: str, right: str):
        left_cfg = cfg["MODEL_CFG"][left]
        right_cfg = cfg["MODEL_CFG"][right]

        a_label = f"tag={left}, vendor={left_cfg['vendor']}, model={left_cfg['model']}"
        b_label = f"tag={right}, vendor={right_cfg['vendor']}, model={right_cfg['model']}"

        judge_user = build_judge_user_content(
            user_input=query,
            expected_case=exp_case,
            a_label=a_label,
            a_output=content_of(left),
            b_label=b_label,
            b_output=content_of(right),
        )

        j_ok, j_payload, j_meta = await limited_call(
            call_model_chat(
                vendor=judge_vendor,
                client=clients["J"],
                base_url=judge_base,
                api_key=judge_key,
                model=judge_model,
                system_prompt=judge_sys,
                user_prompt=judge_user,
                force_json=True
            )
        )

        evaluation_obj = None
        status = "failed"
        if j_ok:
            try:
                evaluation_obj = json.loads(j_payload["content"])
                status = "success"
            except Exception:
                status = "failed"

        # map winner -> winner_tag
        winner_tag = None
        if isinstance(evaluation_obj, dict):
            w = evaluation_obj.get("winner")
            if w == "A":
                winner_tag = left
            elif w == "B":
                winner_tag = right
            elif w == "tie":
                winner_tag = "tie"

        return {
            "left": left,
            "right": right,
            "judge_ok": j_ok,
            "judge_error": None if j_ok else j_payload,
            "evaluation_status": status,
            "evaluation": evaluation_obj,
            "winner_tag": winner_tag,
            "judge_meta": j_meta,
        }

    pair_tasks = [asyncio.create_task(judge_pair(l, r)) for (l, r) in pairs]
    pair_results: List[dict] = []
    for fut in asyncio.as_completed(pair_tasks):
        pr = await fut
        pair_results.append(pr)
        if pair_pbar is not None:
            pair_pbar.update(1)

    # 4) Scoreboard + champion (points: win=1, tie=0.5 each)
    scoreboard: Dict[str, Dict[str, Any]] = {t: {"wins": 0, "losses": 0, "ties": 0, "points": 0.0} for t in model_tags}

    # also aggregate judge usage across all pairs
    judge_agg = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0, "cached_tokens": 0, "duration": 0.0}

    # quick lookup for head-to-head tiebreak
    head2head: Dict[Tuple[str, str], float] = {}  # (t1,t2) -> points t1 got vs t2

    for pr in pair_results:
        l = pr["left"]
        r = pr["right"]
        ev = pr.get("evaluation") or {}
        w = ev.get("winner") if isinstance(ev, dict) else None

        # judge usage sum
        jm = pr.get("judge_meta") or {}
        judge_agg["input_tokens"] += _tok(jm, "input_tokens")
        judge_agg["output_tokens"] += _tok(jm, "output_tokens")
        judge_agg["total_tokens"] += _tok(jm, "total_tokens")
        judge_agg["cached_tokens"] += _tok(jm, "cached_tokens")
        judge_agg["duration"] += _dur(jm)

        if w == "A":
            scoreboard[l]["wins"] += 1
            scoreboard[r]["losses"] += 1
            scoreboard[l]["points"] += 1.0
            head2head[(l, r)] = head2head.get((l, r), 0.0) + 1.0
            head2head[(r, l)] = head2head.get((r, l), 0.0) + 0.0
        elif w == "B":
            scoreboard[r]["wins"] += 1
            scoreboard[l]["losses"] += 1
            scoreboard[r]["points"] += 1.0
            head2head[(r, l)] = head2head.get((r, l), 0.0) + 1.0
            head2head[(l, r)] = head2head.get((l, r), 0.0) + 0.0
        elif w == "tie":
            scoreboard[l]["ties"] += 1
            scoreboard[r]["ties"] += 1
            scoreboard[l]["points"] += 0.5
            scoreboard[r]["points"] += 0.5
            head2head[(l, r)] = head2head.get((l, r), 0.0) + 0.5
            head2head[(r, l)] = head2head.get((r, l), 0.0) + 0.5
        else:
            # judge failed or unknown => no points
            pass

    # champion selection with tie-break: points -> head-to-head among tied -> order in MODELS
    max_points = max(scoreboard[t]["points"] for t in model_tags) if model_tags else 0.0
    top = [t for t in model_tags if scoreboard[t]["points"] == max_points]
    tie_break = "points"
    champion = top[0] if top else None

    if len(top) > 1:
        # head-to-head mini score
        best = None
        best_score = -1.0
        for t in top:
            s = 0.0
            for o in top:
                if o == t:
                    continue
                s += head2head.get((t, o), 0.0)
            if s > best_score:
                best_score = s
                best = t
        if best is not None and best_score > -1.0:
            champion = best
            tie_break = "head_to_head"
        else:
            champion = top[0]
            tie_break = "order"

    tournament = {
        "models": model_tags,
        "pairwise": sorted(pair_results, key=lambda x: (x["left"], x["right"])),
        "scoreboard": scoreboard,
        "champion": {"tag": champion, "points": scoreboard[champion]["points"] if champion else None, "tie_break": tie_break},
    }

    # 5) Primary pair evaluation = pair(tagA, tagB) (legacy "evaluation")
    primary_pair_eval = None
    primary_pair_status = "failed"
    primary_j_ok = False
    primary_j_err = None

    for pr in pair_results:
        if pr["left"] == tagA and pr["right"] == tagB:
            primary_pair_eval = pr.get("evaluation")
            primary_pair_status = pr.get("evaluation_status") or "failed"
            primary_j_ok = bool(pr.get("judge_ok"))
            primary_j_err = pr.get("judge_error")
            break

    # 6) usage_breakdown_by_tag (models + judge aggregated)
    usage_by_tag: Dict[str, Dict[str, Any]] = {}
    for t in model_tags:
        ok, _payload, m = model_results[t]
        usage_by_tag[t] = {
            "input_tokens": _tok(m, "input_tokens"),
            "output_tokens": _tok(m, "output_tokens"),
            "total_tokens": _tok(m, "total_tokens"),
            "cached_tokens": _tok(m, "cached_tokens"),
            "duration": _dur(m),
        }
    usage_by_tag["J"] = dict(judge_agg)

    # legacy usage_breakdown (A/B/J slots) = first two tags + judge
    usage_breakdown = {
        "A": usage_by_tag[tagA],
        "B": usage_by_tag[tagB],
        "J": usage_by_tag["J"],
    }

    total_input_tokens = sum(v["input_tokens"] for v in usage_by_tag.values())
    total_output_tokens = sum(v["output_tokens"] for v in usage_by_tag.values())
    total_tokens = sum(v["total_tokens"] for v in usage_by_tag.values())
    total_duration = sum(v["duration"] for v in usage_by_tag.values())

    record = {
        "query_index": idx,
        "user_input": query,
        "expected_case": exp_case,

        # legacy primary outputs
        "model_a_output": a_content,
        "model_b_output": b_content,

        # legacy evaluation (primary pair)
        "evaluation": primary_pair_eval,
        "evaluation_status": primary_pair_status,
        "timestamp": datetime.now(UTC).isoformat(),

        "api_call_metadata": {
            "duration_seconds": round(total_duration, 2),
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "total_tokens": total_tokens
        },

        "usage_breakdown": usage_breakdown,

        # ✅ 확장(새): 태그 전체 breakdown / 출력 / 토너먼트
        "model_a_tag": tagA,
        "model_b_tag": tagB,
        "usage_breakdown_by_tag": usage_by_tag,
        "model_outputs_by_tag": {t: content_of(t) for t in model_tags},
        "tournament": tournament,

        # legacy model identifiers (primary pair)
        "model_a": cfg["MODEL_CFG"][tagA]["model"],
        "model_b": cfg["MODEL_CFG"][tagB]["model"],

        "a_ok": a_ok,
        "b_ok": b_ok,
        "a_error": None if a_ok else a_payload,
        "b_error": None if b_ok else b_payload,
        "judge_ok": primary_j_ok,
        "judge_error": None if primary_j_ok else primary_j_err,

        # legacy parsing/validation for primary pair
        "a_parse_ok": a_parsed is not None,
        "b_parse_ok": b_parsed is not None,
        "model_a_case": a_case,
        "model_b_case": b_case,

        "a_function_name": a_c2.get("function_name"),
        "a_available_function": a_c2.get("available_function"),
        "a_function_in_whitelist": a_c2.get("function_in_whitelist"),
        "a_available_in_whitelist": a_c2.get("available_in_whitelist"),

        "b_function_name": b_c2.get("function_name"),
        "b_available_function": b_c2.get("available_function"),
        "b_function_in_whitelist": b_c2.get("function_in_whitelist"),
        "b_available_in_whitelist": b_c2.get("available_in_whitelist"),

        "a_prepared_number": a_c3.get("prepared_number"),
        "a_prepared_in_range": a_c3.get("prepared_in_range"),
        "b_prepared_number": b_c3.get("prepared_number"),
        "b_prepared_in_range": b_c3.get("prepared_in_range"),

        "expected_function": exp_fn,
        "expected_available_function": exp_act,
        "expected_prepared_question": exp_pq,

        "a_function_matches_expected": a_func_eq,
        "a_action_matches_expected": a_act_eq,
        "b_function_matches_expected": b_func_eq,
        "b_action_matches_expected": b_act_eq,
        "a_pq_matches_expected": a_pq_eq,
        "b_pq_matches_expected": b_pq_eq,
    }
    return record


async def main_async(args):
    cfg = load_env()
    model_tags: List[str] = cfg["MODELS"]

    # sys prompts 로드(ENV 기본 + CLI override)
    sys_prompts: Dict[str, str] = {}
    for t in model_tags:
        env_path = cfg["MODEL_CFG"][t].get("sys_prompt_path") or ""
        override = ""
        if t == "A" and args.model_a_prompt:
            override = args.model_a_prompt
        elif t == "B" and args.model_b_prompt:
            override = args.model_b_prompt
        elif t == "C" and args.model_c_prompt:
            override = args.model_c_prompt

        use_path = override or env_path
        if not use_path:
            raise ValueError(f"[PROMPT] Missing sys prompt path for tag {t}. Set MODEL_{t}_SYS_PROMPT or pass CLI override.")
        sys_prompts[t] = _read_text_file(use_path)

    judge_sys = _read_text_file(args.judge_prompt)

    func_spec = None
    if args.functions_spec:
        p = Path(args.functions_spec)
        if p.exists():
            func_spec = json.loads(p.read_text(encoding="utf-8"))

    # read tests
    rows = []
    with open(args.tests, "r", encoding="utf-8") as f:
        rdr = csv.DictReader(f)
        for r in rdr:
            rows.append(r)

    out_dir = Path(args.out)
    jsonl_path, xlsx_path = make_dated_output_paths(out_dir, base_prefix="results")

    # progress bars
    n = len(model_tags)
    pairs_count = (n * (n - 1)) // 2
    pair_pbar = tqdm(total=len(rows) * pairs_count, desc="Judging", unit="match")

    sema = asyncio.Semaphore(cfg["CONC"])

    async with AsyncExitStack() as stack:
        clients: Dict[str, httpx.AsyncClient] = {}
        for t in model_tags:
            mc = cfg["MODEL_CFG"][t]
            clients[t] = await stack.enter_async_context(
                make_client(mc["vendor"], mc["base_url"], mc["api_key"], cfg["TIMEOUT"])
            )
        # judge client
        j = cfg["JUDGE"]
        clients["J"] = await stack.enter_async_context(
            make_client(j["vendor"], j["base_url"], j["api_key"], cfg["TIMEOUT"])
        )

        recs: List[dict] = []
        with jsonl_path.open("w", encoding="utf-8") as fw:
            for i, row in enumerate(tqdm(rows, desc="Evaluating", unit="q")):
                try:
                    rec = await process_one(
                        idx=i,
                        row=row,
                        cfg=cfg,
                        clients=clients,
                        sys_prompts=sys_prompts,
                        judge_sys=judge_sys,
                        func_spec=func_spec,
                        prepared_min=args.prepared_min,
                        prepared_max=args.prepared_max,
                        sema=sema,
                        pair_pbar=pair_pbar,
                    )
                except Exception as e:
                    print(f"\n[ABORT] Unexpected runner error at row {i}: {e!r}")
                    break

                fw.write(json.dumps(rec, ensure_ascii=False) + "\n")
                recs.append(rec)

    pair_pbar.close()

    summary = build_summary(recs, cfg["COST"], model_tags)
    write_excel_from_jsonl(jsonl_path, xlsx_path, summary)

    print(f"\nDone.\n- JSONL: {jsonl_path}\n- Excel : {xlsx_path}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--tests", type=str, required=True,
                        help="CSV file with columns: query, expected_case, expected_function, expected_available_function, expected_prepared_question")
    # legacy override (optional)
    parser.add_argument("--model-a-prompt", type=str, default="", help="Override for A sys prompt (optional)")
    parser.add_argument("--model-b-prompt", type=str, default="", help="Override for B sys prompt (optional)")
    parser.add_argument("--model-c-prompt", type=str, default="", help="Override for C sys prompt (optional)")
    parser.add_argument("--judge-prompt", type=str, required=True)
    parser.add_argument("--functions-spec", type=str, default="", help="JSON mapping of function -> [allowed actions]")
    parser.add_argument("--prepared-min", type=int, default=1)
    parser.add_argument("--prepared-max", type=int, default=60)
    parser.add_argument("--out", type=str, required=True, help="Output directory")
    asyncio.run(main_async(parser.parse_args()))